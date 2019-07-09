#-*- coding: utf-8 -*-

from konlpy.tag import Okt
from konlpy.tag import Kkma
from openpyxl import load_workbook
import re



from nltk.tokenize import word_tokenize, sent_tokenize
import math

okt = Okt()
load_wb = load_workbook("Samsung.xlsx", data_only=True)
load_ws = load_wb['sheet1']


def cleanText(rawText):
    toBeTerminated = '[-=+;,↓#/\↑?▼:★^△▶’●▲”$.·∼@◆*···\"※~&ⓒ%ㆍ■!』\\‘|\(\)\[\]\<\>`\'…》]'
    strippedText = re.sub(toBeTerminated,'',rawText)
    return strippedText

wordDict = {}

#셀 좌표로 값 출력
for i in range(2,450):
    kkma = Kkma()
    titles = load_ws.cell(i,3).value
    titles = cleanText(titles)

    contents = load_ws.cell(i,5).value
    contents = cleanText(contents)
    

    texts = titles + contents

    malist = kkma.pos(texts)
   
    for x in malist:
        if 'NNG' in x :
            keyword = x[0]
            if keyword in wordDict:
                wordDict[keyword] += 1
            else:
                wordDict[keyword] = 1

wordDict = sorted(wordDict.items(), key=lambda t : t[1], reverse=True)

# print(wordDict)

#####################
text1 = ""
for i in range(2,450):
    titles = load_ws.cell(i,3).value
    titles = cleanText(titles)

    contents = load_ws.cell(i,5).value
    contents = cleanText(contents)
    
    text1 += (titles + contents + "\n")
# print(text1)

def get_doc(sents) :
    """
    이 함수는 텍스트를 문장으로 나누고 각 문장을 document로 
    여겨서 각 document의 총 단어 등장 수를 카운트
    """
    doc_info = []
    i = 0
    for x in sents:
        i += 1
        count = count_words(sents)
        temp = {'doc_id' : i, 'doc_length' : count}
        doc_info.append(temp)
    return doc_info

def count_words(sent) :
    """
    이 함수는 각 input 텍스트의 총 단어수를 리턴
    """
    count = 0
    words = krword_tokenize(sent)
    for word in words :
        count += 1
    return count

def krword_tokenize(sent):
    result = []
    sample = okt.pos(sent, norm=True, stem=True)
    for hts in sample:
        if 'Noun' in hts or 'alpha' in hts:
            result.append(hts)
    return result

def create_freq_dict(sents) :
    """
    이 함수는 각 사전의 단어들에 대한 빈도수를 dictionary로 만듬
    """
    i = 0
    freqDict_list = []
    for sent in sents :
        i += 1
        freq_dict = {}
        words = word_tokenize(sent)
        for word in words :
            word = word.lower()
            if word in freq_dict :
                freq_dict[word] += 1
            else :
                freq_dict[word] = 1
            temp = {'doc_id' : i, 'freq_dict' : freq_dict}
        freqDict_list.append(temp)
    return freqDict_list

def computeTF(doc_info, freqDict_list) :
    """
    tf = (document 내에서 단어 등장 빈도 / document내의 전체 단어 수)
    """
    
    TF_scores = []
    for tempDict in freqDict_list :
        id = tempDict['doc_id']
        for k in tempDict['freq_dict'] :
          
            temp = {'doc_id' : id,
                   'TF_score' : tempDict['freq_dict'][k]/doc_info[id-1]['doc_length'],
                   'key' : k}
            TF_scores.append(temp)
        return TF_scores

def computeIDF(doc_info, freqDict_list) :
    """
    idf = log(전체 문서 수/ 단어가 드 안에 들어있는 문서 수)
    """
    IDF_scores = []
    counter = 0
    for dict in freqDict_list :
        counter += 1
        for k in dict['freq_dict'].keys() :
            count = sum([k in tempDict['freq_dict'] for tempDict in freqDict_list])
            temp = {'doc_id' : counter, 'IDF_score' : math.log(len(doc_info)/count), 'key' : k}
            IDF_scores.append(temp)
        
    return IDF_scores

def computeTFIDF(TF_scores, IDF_scores) :
    TFIDF_scores = []
    for j in IDF_scores :
        for i in TF_scores :
            if j['key'] == i['key'] and j ['doc_id'] == i['doc_id'] :
                temp = {'doc_id' : j['doc_id'],
                       'TFIDF_score' : j['IDF_score']*i['TF_score'],
                       'key' : i['key']}
        TFIDF_scores.append(temp)
    return TFIDF_scores

def remove_string_special_characters(s) :
    """
    s : 문자열
    """
    stripped = re.sub('[^\w\s]', '', s)
    #문자열 중간 공백 삭제
    stripped = re.sub('\s+', ' ', stripped)
    #처음과 끝 공백 삭제
    stripped = stripped.strip()

    return stripped

def onlyNouns(article):
    result = ""
    sample = okt.pos(article, norm=True, stem=True)
    for hts in sample:
        print(hts)
        if 'Noun' in hts or 'alpha' in hts:
            result += (hts[0]) + ' '
    return result


x = text1.split('\n')

doc_info = get_doc(x)

print(doc_info)

# freqDict_list = create_freq_dict(x)
# TF_scores = computeTF(doc_info, freqDict_list)
# IDF_scores = computeIDF(doc_info, freqDict_list)
# print("TF_scores are",TF_scores)
# print("IDF_scroes are",IDF_scores)
#
# doc_info
#
# freqDict_list